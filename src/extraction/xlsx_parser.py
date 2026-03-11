"""Excel file parsing for TBT, E&I Compliance, and generic spreadsheets.

Supports:
- .xlsx via openpyxl
- .xls  via xlrd (legacy Excel format)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


# ── Structured models ────────────────────────────────────────────────────────


class TBTItem(BaseModel):
    """One row from a Technical Bid Evaluation Table (TBT).

    Status codes:
        F = Fully compliant
        A = Acceptable (with comments)
        X = Not compliant
        O = To be confirmed at detailed engineering
        Y = Not applicable
        C = Clarification required
        E = To be confirmed at detailed engineering
    """

    row_number: int
    section: str = ""
    description: str = ""
    spec_requirement: str = ""
    bidder_response: str = ""
    status: str = ""
    remarks: str = ""


class ComplianceCheckItem(BaseModel):
    """One row from the E&I Technical Compliance Check Sheet."""

    row_number: int
    section_no: str = ""
    subject: str = ""
    technical_requirement: str = ""
    project_requirement: str = ""
    relevant_document: str = ""
    complies: bool | None = None
    remarks: str = ""


class ExcelSheet(BaseModel):
    """Generic representation of a parsed Excel sheet."""

    sheet_name: str
    headers: list[str] = Field(default_factory=list)
    rows: list[dict[str, Any]] = Field(default_factory=list)
    total_rows: int = 0


# ── Parser ───────────────────────────────────────────────────────────────────


class ExcelParser:
    """Parse Excel files into structured data."""

    def parse_tbt(self, file_path: str | Path) -> list[TBTItem]:
        """Parse a Technical Bid Evaluation Table (.xlsx).

        Expects columns roughly like:
        Section | Description | Spec Requirement | Bidder Response | Status | Remarks

        Returns
        -------
        list[TBTItem]
            Parsed rows, skipping empty rows and headers.
        """
        path = Path(file_path)
        self._check_file(path)

        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError(f"No active sheet in {path.name}")

        items: list[TBTItem] = []
        header_row = self._find_header_row(ws)

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            cells = [str(c).strip() if c is not None else "" for c in row]

            # TBT typically has many columns; we take the most relevant ones.
            # Column mapping depends on actual file — we use a flexible approach.
            item = TBTItem(
                row_number=row_idx,
                section=cells[0] if len(cells) > 0 else "",
                description=cells[1] if len(cells) > 1 else "",
                spec_requirement=cells[2] if len(cells) > 2 else "",
                bidder_response=self._find_bidder_column(cells),
                status=self._find_status(cells),
                remarks=cells[-1] if len(cells) > 3 else "",
            )
            items.append(item)

        wb.close()
        logger.info("Parsed TBT %s: %d items", path.name, len(items))
        return items

    def parse_compliance_check(self, file_path: str | Path) -> list[ComplianceCheckItem]:
        """Parse an E&I Technical Compliance Check Sheet.

        Supports both .xlsx and .xls formats.

        Returns
        -------
        list[ComplianceCheckItem]
            Parsed compliance items.
        """
        path = Path(file_path)
        self._check_file(path)

        if path.suffix.lower() == ".xls":
            return self._parse_compliance_xls(path)
        return self._parse_compliance_xlsx(path)

    def parse_generic(self, file_path: str | Path) -> list[ExcelSheet]:
        """Parse any Excel file into a list of sheets with rows as dicts.

        Returns
        -------
        list[ExcelSheet]
            One entry per sheet with headers and row dicts.
        """
        path = Path(file_path)
        self._check_file(path)

        if path.suffix.lower() == ".xls":
            return self._parse_generic_xls(path)
        return self._parse_generic_xlsx(path)

    # ── Private helpers ──────────────────────────────────────────────────────

    @staticmethod
    def _check_file(path: Path) -> None:
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"Not an Excel file: {path}")

    @staticmethod
    def _find_header_row(ws: Any, max_scan: int = 15) -> int:
        """Find the header row by looking for the first row with >=3 non-empty cells."""
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True), start=1):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 3:
                return row_idx
        return 1

    @staticmethod
    def _find_bidder_column(cells: list[str]) -> str:
        """Heuristic: bidder response is often in the middle-right of the row."""
        if len(cells) >= 6:
            return cells[3]
        return cells[-2] if len(cells) >= 4 else ""

    @staticmethod
    def _find_status(cells: list[str]) -> str:
        """Look for a typical status code (F/A/X/O/Y/C/E) in the row."""
        status_codes = {"F", "A", "X", "O", "Y", "C", "E"}
        for cell in reversed(cells):
            if cell.strip().upper() in status_codes:
                return cell.strip().upper()
        return ""

    def _parse_compliance_xlsx(self, path: Path) -> list[ComplianceCheckItem]:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError(f"No active sheet in {path.name}")

        items: list[ComplianceCheckItem] = []
        header_row = self._find_header_row(ws)

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            cells = [str(c).strip() if c is not None else "" for c in row]
            complies = self._parse_complies(cells)

            items.append(
                ComplianceCheckItem(
                    row_number=row_idx,
                    section_no=cells[0] if len(cells) > 0 else "",
                    subject=cells[1] if len(cells) > 1 else "",
                    technical_requirement=cells[2] if len(cells) > 2 else "",
                    project_requirement=cells[3] if len(cells) > 3 else "",
                    relevant_document=cells[4] if len(cells) > 4 else "",
                    complies=complies,
                    remarks=cells[-1] if len(cells) > 5 else "",
                )
            )

        wb.close()
        logger.info("Parsed compliance XLSX %s: %d items", path.name, len(items))
        return items

    def _parse_compliance_xls(self, path: Path) -> list[ComplianceCheckItem]:
        import xlrd

        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)

        items: list[ComplianceCheckItem] = []
        header_row = self._find_header_row_xls(ws)

        for row_idx in range(header_row, ws.nrows):
            cells = [str(ws.cell_value(row_idx, col)).strip() for col in range(ws.ncols)]
            if all(c == "" for c in cells):
                continue

            complies = self._parse_complies(cells)

            items.append(
                ComplianceCheckItem(
                    row_number=row_idx + 1,
                    section_no=cells[0] if len(cells) > 0 else "",
                    subject=cells[1] if len(cells) > 1 else "",
                    technical_requirement=cells[2] if len(cells) > 2 else "",
                    project_requirement=cells[3] if len(cells) > 3 else "",
                    relevant_document=cells[4] if len(cells) > 4 else "",
                    complies=complies,
                    remarks=cells[-1] if len(cells) > 5 else "",
                )
            )

        logger.info("Parsed compliance XLS %s: %d items", path.name, len(items))
        return items

    @staticmethod
    def _find_header_row_xls(ws: Any, max_scan: int = 15) -> int:
        for row_idx in range(min(max_scan, ws.nrows)):
            non_empty = sum(1 for col in range(ws.ncols) if str(ws.cell_value(row_idx, col)).strip())
            if non_empty >= 3:
                return row_idx + 1  # skip header itself
        return 1

    @staticmethod
    def _parse_complies(cells: list[str]) -> bool | None:
        """Parse Yes/No/Y/N compliance value from row cells."""
        yes_no = {"yes", "y", "si", "sí"}
        no_vals = {"no", "n"}
        for cell in cells:
            lower = cell.lower().strip()
            if lower in yes_no:
                return True
            if lower in no_vals:
                return False
        return None

    def _parse_generic_xlsx(self, path: Path) -> list[ExcelSheet]:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets: list[ExcelSheet] = []

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                sheets.append(ExcelSheet(sheet_name=ws_name))
                continue

            # First non-empty row = headers
            headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(all_rows[0])]
            rows: list[dict[str, Any]] = []
            for row in all_rows[1:]:
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                row_dict = {
                    headers[i]: (str(c).strip() if c is not None else "")
                    for i, c in enumerate(row)
                    if i < len(headers)
                }
                rows.append(row_dict)

            sheets.append(
                ExcelSheet(
                    sheet_name=ws_name,
                    headers=headers,
                    rows=rows,
                    total_rows=len(rows),
                )
            )

        wb.close()
        return sheets

    def _parse_generic_xls(self, path: Path) -> list[ExcelSheet]:
        import xlrd

        wb = xlrd.open_workbook(str(path))
        sheets: list[ExcelSheet] = []

        for ws_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(ws_idx)
            if ws.nrows == 0:
                sheets.append(ExcelSheet(sheet_name=ws.name))
                continue

            headers = [str(ws.cell_value(0, col)).strip() or f"col_{col}" for col in range(ws.ncols)]
            rows: list[dict[str, Any]] = []
            for row_idx in range(1, ws.nrows):
                cells = {
                    headers[col]: str(ws.cell_value(row_idx, col)).strip()
                    for col in range(ws.ncols)
                    if col < len(headers)
                }
                if all(v == "" for v in cells.values()):
                    continue
                rows.append(cells)

            sheets.append(
                ExcelSheet(
                    sheet_name=ws.name,
                    headers=headers,
                    rows=rows,
                    total_rows=len(rows),
                )
            )

        return sheets
